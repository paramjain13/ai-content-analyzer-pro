from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
import re

def extract_text_from_xlsx(file_path, max_rows=100):
    """
    Extract text and data from XLSX file
    Returns: (text, metadata)
    """
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        
        # Extract metadata
        metadata = {
            "title": wb.properties.title or "Excel Spreadsheet",
            "author": wb.properties.creator or "Unknown",
            "created": str(wb.properties.created) if wb.properties.created else "Unknown",
            "modified": str(wb.properties.modified) if wb.properties.modified else "Unknown",
            "sheets": len(wb.sheetnames),
            "sheet_names": wb.sheetnames
        }
        
        text_parts = []
        
        # Process each worksheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Add sheet name as section header
            text_parts.append(f"[Sheet: {sheet_name}]")
            
            # Get dimensions
            max_row = min(ws.max_row, max_rows) if ws.max_row else max_rows
            
            # Extract data from cells
            rows_processed = 0
            for row_num, row in enumerate(ws.iter_rows(max_row=max_row, values_only=True), 1):
                if rows_processed >= max_rows:
                    break
                
                # Convert row to text
                row_data = []
                for cell_value in row:
                    if cell_value is not None:
                        # Convert to string and clean
                        cell_str = str(cell_value).strip()
                        if cell_str:
                            row_data.append(cell_str)
                
                if row_data:
                    # Join cell values with commas
                    row_text = ", ".join(row_data)
                    text_parts.append(row_text)
                    rows_processed += 1
        
        wb.close()
        
        full_text = " ".join(text_parts)
        
        # Clean up whitespace
        full_text = re.sub(r'\s+', ' ', full_text).strip()
        
        if not full_text or len(full_text) < 50:
            raise Exception("No readable data found in XLSX file")
        
        return full_text, metadata
        
    except InvalidFileException:
        raise Exception("Invalid or corrupted XLSX file")
    except Exception as e:
        raise Exception(f"Error reading XLSX file: {str(e)}")